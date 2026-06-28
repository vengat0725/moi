from flask import Flask, request, jsonify, send_file, render_template
import sqlite3
import os
from datetime import datetime
from io import BytesIO

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "moi_tracker.db")


# ---------------------------------------------------------------------------
# Database setup
# ---------------------------------------------------------------------------
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS moi_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            guest_name TEXT NOT NULL,
            amount REAL NOT NULL,
            relation TEXT,
            event_name TEXT,
            entry_date TEXT NOT NULL,
            notes TEXT,
            created_at TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()


init_db()


# ---------------------------------------------------------------------------
# Page
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# API: list / search entries
# ---------------------------------------------------------------------------
@app.route("/api/entries", methods=["GET"])
def get_entries():
    search = request.args.get("search", "").strip()
    sort_by = request.args.get("sort", "date_desc")

    conn = get_db()
    query = "SELECT * FROM moi_entries WHERE 1=1"
    params = []

    if search:
        query += " AND (guest_name LIKE ? OR relation LIKE ? OR event_name LIKE ?)"
        like = f"%{search}%"
        params.extend([like, like, like])

    sort_map = {
        "date_desc": "entry_date DESC, id DESC",
        "date_asc": "entry_date ASC, id ASC",
        "amount_desc": "amount DESC",
        "amount_asc": "amount ASC",
        "name_asc": "guest_name ASC",
    }
    query += f" ORDER BY {sort_map.get(sort_by, sort_map['date_desc'])}"

    rows = conn.execute(query, params).fetchall()

    total_row = conn.execute(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(amount),0) as total FROM moi_entries"
    ).fetchone()
    conn.close()

    entries = [dict(r) for r in rows]
    return jsonify({
        "entries": entries,
        "count": total_row["cnt"],
        "total": total_row["total"],
    })


# ---------------------------------------------------------------------------
# API: create entry
# ---------------------------------------------------------------------------
@app.route("/api/entries", methods=["POST"])
def add_entry():
    data = request.get_json(force=True)

    guest_name = (data.get("guest_name") or "").strip()
    amount = data.get("amount")
    relation = (data.get("relation") or "").strip()
    event_name = (data.get("event_name") or "").strip()
    entry_date = (data.get("entry_date") or "").strip()
    notes = (data.get("notes") or "").strip()

    if not guest_name:
        return jsonify({"error": "Guest name is required"}), 400
    try:
        amount = float(amount)
        if amount < 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Amount must be a valid positive number"}), 400
    if not entry_date:
        entry_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_db()
    cur = conn.execute(
        """INSERT INTO moi_entries
           (guest_name, amount, relation, event_name, entry_date, notes, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (guest_name, amount, relation, event_name, entry_date, notes,
         datetime.now().isoformat()),
    )
    conn.commit()
    new_id = cur.lastrowid
    row = conn.execute("SELECT * FROM moi_entries WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


# ---------------------------------------------------------------------------
# API: update entry
# ---------------------------------------------------------------------------
@app.route("/api/entries/<int:entry_id>", methods=["PUT"])
def update_entry(entry_id):
    data = request.get_json(force=True)

    guest_name = (data.get("guest_name") or "").strip()
    amount = data.get("amount")
    relation = (data.get("relation") or "").strip()
    event_name = (data.get("event_name") or "").strip()
    entry_date = (data.get("entry_date") or "").strip()
    notes = (data.get("notes") or "").strip()

    if not guest_name:
        return jsonify({"error": "Guest name is required"}), 400
    try:
        amount = float(amount)
        if amount < 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Amount must be a valid positive number"}), 400

    conn = get_db()
    existing = conn.execute("SELECT id FROM moi_entries WHERE id=?", (entry_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "Entry not found"}), 404

    conn.execute(
        """UPDATE moi_entries
           SET guest_name=?, amount=?, relation=?, event_name=?, entry_date=?, notes=?
           WHERE id=?""",
        (guest_name, amount, relation, event_name, entry_date, notes, entry_id),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM moi_entries WHERE id=?", (entry_id,)).fetchone()
    conn.close()
    return jsonify(dict(row))


# ---------------------------------------------------------------------------
# API: delete entry
# ---------------------------------------------------------------------------
@app.route("/api/entries/<int:entry_id>", methods=["DELETE"])
def delete_entry(entry_id):
    conn = get_db()
    existing = conn.execute("SELECT id FROM moi_entries WHERE id=?", (entry_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "Entry not found"}), 404
    conn.execute("DELETE FROM moi_entries WHERE id=?", (entry_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Export: Excel
# ---------------------------------------------------------------------------
@app.route("/api/export/excel", methods=["GET"])
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM moi_entries ORDER BY entry_date ASC, id ASC"
    ).fetchall()
    total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) as total FROM moi_entries"
    ).fetchone()["total"]
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Moi Register"

    maroon = "8B2635"
    gold = "C7912B"
    ivory = "FBF6EC"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Moi Register"
    ws["A1"].font = Font(name="Georgia", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=maroon)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Date", "Guest Name", "Relation", "Event", "Amount (₹)", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=gold)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D8CBB0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 3
    for row in rows:
        values = [row["entry_date"], row["guest_name"], row["relation"] or "-",
                   row["event_name"] or "-", row["amount"], row["notes"] or ""]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            if col == 5:
                cell.number_format = "₹ #,##0.00"
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ivory)
        r += 1

    r += 1
    ws.cell(row=r, column=4, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=r, column=5, value=total)
    total_cell.font = Font(bold=True, color=maroon)
    total_cell.number_format = "₹ #,##0.00"

    widths = [14, 22, 16, 20, 16, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"moi_register_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Export: PDF
# ---------------------------------------------------------------------------
@app.route("/api/export/pdf", methods=["GET"])
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM moi_entries ORDER BY entry_date ASC, id ASC"
    ).fetchall()
    total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) as total, COUNT(*) as cnt FROM moi_entries"
    ).fetchone()
    conn.close()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             topMargin=18 * mm, bottomMargin=18 * mm,
                             leftMargin=14 * mm, rightMargin=14 * mm)

    maroon = colors.HexColor("#8B2635")
    gold = colors.HexColor("#C7912B")
    ivory = colors.HexColor("#FBF6EC")
    charcoal = colors.HexColor("#2B2520")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  textColor=maroon, fontSize=22,
                                  alignment=TA_CENTER, spaceAfter=2)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                                textColor=charcoal, fontSize=10,
                                alignment=TA_CENTER, spaceAfter=14)

    elements = [
        Paragraph("Moi Register", title_style),
        Paragraph(f"Generated on {datetime.now().strftime('%d %b %Y')} &nbsp;|&nbsp; "
                   f"{total['cnt']} entries", sub_style),
    ]

    data = [["Date", "Guest Name", "Relation", "Event", "Amount (₹)"]]
    for row in rows:
        data.append([
            row["entry_date"], row["guest_name"], row["relation"] or "-",
            row["event_name"] or "-", f"{row['amount']:,.2f}",
        ])
    data.append(["", "", "", "TOTAL", f"{total['total']:,.2f}"])

    table = Table(data, colWidths=[24 * mm, 42 * mm, 30 * mm, 38 * mm, 30 * mm], repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), gold),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -2), 0.4, colors.HexColor("#D8CBB0")),
        ("FONTSIZE", (0, 1), (-1, -1), 9.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, ivory]),
        ("LINEABOVE", (0, -1), (-1, -1), 1, maroon),
        ("FONTNAME", (3, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (3, -1), (-1, -1), maroon),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    filename = f"moi_register_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype="application/pdf")


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000)
